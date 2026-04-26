#!/usr/bin/env python3
"""
Rebuild catalog data from source Excel files.

Usage:
  python scripts/rebuild_catalog.py \
    --excel-dir "C:/Users/user/Desktop/Work/Черновик/Mansvalve"

The script generates:
  - data/catalog-products.json
  - lib/catalog-data.ts
"""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


CATEGORIES: list[dict[str, Any]] = [
    {
        "id": "zadvizhki",
        "name": "Задвижки",
        "slug": "zadvizhki",
        "subcategories": [
            {
                "id": "zadvizhki-chugunnye",
                "name": "Чугунные задвижки",
                "slug": "zadvizhki-chugunnye",
                "parentCategory": "zadvizhki",
            },
            {
                "id": "zadvizhki-stalnyye",
                "name": "Стальные задвижки",
                "slug": "zadvizhki-stalnyye",
                "parentCategory": "zadvizhki",
            },
            {
                "id": "zadvizhki-shibernye",
                "name": "Шиберные (ножевые) задвижки",
                "slug": "zadvizhki-shibernye",
                "parentCategory": "zadvizhki",
            },
            {
                "id": "zadvizhki-klinovye",
                "name": "Клиновые задвижки",
                "slug": "zadvizhki-klinovye",
                "parentCategory": "zadvizhki",
            },
            {
                "id": "zadvizhki-s-elektroprivodom",
                "name": "Задвижки с электроприводом",
                "slug": "zadvizhki-s-elektroprivodom",
                "parentCategory": "zadvizhki",
            },
            {
                "id": "zadvizhki-pn40-pn64",
                "name": "Задвижки PN40 / PN64",
                "slug": "zadvizhki-pn40-pn64",
                "parentCategory": "zadvizhki",
            },
        ],
    },
    {
        "id": "krany-sharovye",
        "name": "Краны шаровые",
        "slug": "krany-sharovye",
        "subcategories": [
            {
                "id": "krany-flantsevy",
                "name": "Фланцевые",
                "slug": "krany-flantsevy",
                "parentCategory": "krany-sharovye",
            },
            {
                "id": "krany-pod-privarku",
                "name": "Под приварку",
                "slug": "krany-pod-privarku",
                "parentCategory": "krany-sharovye",
            },
            {
                "id": "krany-tselnosvarnye",
                "name": "Цельносварные",
                "slug": "krany-tselnosvarnye",
                "parentCategory": "krany-sharovye",
            },
        ],
    },
    {
        "id": "zatvory",
        "name": "Затворы",
        "slug": "zatvory",
        "subcategories": [
            {
                "id": "zatvory-mezhflantsevy",
                "name": "Межфланцевые затворы",
                "slug": "zatvory-mezhflantsevy",
                "parentCategory": "zatvory",
            },
            {
                "id": "zatvory-flantsevy",
                "name": "Фланцевые затворы",
                "slug": "zatvory-flantsevy",
                "parentCategory": "zatvory",
            },
            {
                "id": "zatvory-diskovye",
                "name": "Дисковые поворотные затворы",
                "slug": "zatvory-diskovye",
                "parentCategory": "zatvory",
            },
            {
                "id": "zatvory-s-uplotneniem",
                "name": "Затворы с уплотнением",
                "slug": "zatvory-s-uplotneniem",
                "parentCategory": "zatvory",
            },
        ],
    },
    {
        "id": "klapany",
        "name": "Клапаны",
        "slug": "klapany",
        "subcategories": [
            {
                "id": "klapany-obratnye",
                "name": "Обратные клапаны",
                "slug": "klapany-obratnye",
                "parentCategory": "klapany",
            },
            {
                "id": "klapany-povorotnye",
                "name": "Поворотные клапаны",
                "slug": "klapany-povorotnye",
                "parentCategory": "klapany",
            },
            {
                "id": "klapany-chugunnye",
                "name": "Чугунные водопроводные клапаны",
                "slug": "klapany-chugunnye",
                "parentCategory": "klapany",
            },
        ],
    },
    {
        "id": "filtry-i-kompensatory",
        "name": "Фильтры и компенсаторы",
        "slug": "filtry-i-kompensatory",
        "subcategories": [
            {
                "id": "filtry-korzinchatye",
                "name": "Фильтры корзинчатые",
                "slug": "filtry-korzinchatye",
                "parentCategory": "filtry-i-kompensatory",
            },
            {
                "id": "kompensatory",
                "name": "Компенсаторы",
                "slug": "kompensatory",
                "parentCategory": "filtry-i-kompensatory",
            },
        ],
    },
    {
        "id": "flansy-i-otvody",
        "name": "Фланцы и отводы",
        "slug": "flansy-i-otvody",
        "subcategories": [
            {
                "id": "flansy",
                "name": "Фланцы",
                "slug": "flansy",
                "parentCategory": "flansy-i-otvody",
            },
            {
                "id": "otvody",
                "name": "Отводы",
                "slug": "otvody",
                "parentCategory": "flansy-i-otvody",
            },
        ],
    },
    {
        "id": "krepezh-i-prokladki",
        "name": "Крепёж и прокладки",
        "slug": "krepezh-i-prokladki",
        "subcategories": [
            {
                "id": "krepezh-bolty",
                "name": "Болты",
                "slug": "krepezh-bolty",
                "parentCategory": "krepezh-i-prokladki",
            },
            {
                "id": "krepezh-shpilki",
                "name": "Шпильки",
                "slug": "krepezh-shpilki",
                "parentCategory": "krepezh-i-prokladki",
            },
            {
                "id": "krepezh-gayki",
                "name": "Гайки",
                "slug": "krepezh-gayki",
                "parentCategory": "krepezh-i-prokladki",
            },
            {
                "id": "prokladki-paronit-ponb",
                "name": "Паронит (ПОН-Б)",
                "slug": "prokladki-paronit-ponb",
                "parentCategory": "krepezh-i-prokladki",
            },
        ],
    },
    {
        "id": "elektroprivody",
        "name": "Электроприводы",
        "slug": "elektroprivody",
        "subcategories": [
            {
                "id": "elektroprivody-dlya-zadvizhek",
                "name": "Для задвижек",
                "slug": "elektroprivody-dlya-zadvizhek",
                "parentCategory": "elektroprivody",
            },
            {
                "id": "elektroprivody-dlya-zatvorov",
                "name": "Для затворов",
                "slug": "elektroprivody-dlya-zatvorov",
                "parentCategory": "elektroprivody",
            },
            {
                "id": "elektroprivody-dlya-klapanov",
                "name": "Для клапанов",
                "slug": "elektroprivody-dlya-klapanov",
                "parentCategory": "elektroprivody",
            },
        ],
    },
]


SUBCATEGORY_BY_ID: dict[str, dict[str, str]] = {
    sub["id"]: {"name": sub["name"], "parentCategory": sub["parentCategory"]}
    for category in CATEGORIES
    for sub in category["subcategories"]
}

CATEGORY_NAME_BY_ID: dict[str, str] = {category["id"]: category["name"] for category in CATEGORIES}


CYR_TO_LAT = {
    "а": "a",
    "б": "b",
    "в": "v",
    "г": "g",
    "д": "d",
    "е": "e",
    "ё": "e",
    "ж": "zh",
    "з": "z",
    "и": "i",
    "й": "y",
    "к": "k",
    "л": "l",
    "м": "m",
    "н": "n",
    "о": "o",
    "п": "p",
    "р": "r",
    "с": "s",
    "т": "t",
    "у": "u",
    "ф": "f",
    "х": "h",
    "ц": "ts",
    "ч": "ch",
    "ш": "sh",
    "щ": "sch",
    "ъ": "",
    "ы": "y",
    "ь": "",
    "э": "e",
    "ю": "yu",
    "я": "ya",
}


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower().replace("ё", "е"))


def to_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).strip().replace(" ", "").replace(",", ".")
    if not raw:
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def extract_dn(name: str) -> int | None:
    patterns = [
        r"\bdn\s*[-:]?\s*(\d{1,4})\b",
        r"\bд[уy]\s*[-:]?\s*(\d{1,4})\b",
    ]
    text = normalize_text(name)
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return int(match.group(1))
    return None


def extract_pn(name: str) -> int | None:
    patterns = [
        r"\bpn\s*[-:]?\s*(\d{1,3})\b",
        r"\bр[уy]\s*[-:]?\s*(\d{1,3})\b",
    ]
    text = normalize_text(name)
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return int(match.group(1))
    return None


def extract_thread(name: str) -> str | None:
    text = normalize_text(name)
    match = re.search(r"\b(?:m|м)\s*[-/]?\s*(\d{2})\b", text, flags=re.IGNORECASE)
    if not match:
        return None
    size = int(match.group(1))
    if size < 8 or size > 64:
        return None
    return f"M{size}"


def extract_model(name: str) -> str:
    text = name.strip()
    patterns = [
        r"\b(30[чcс][0-9]{1,3}[а-яa-z]{0,4})\b",
        r"\b(змс\s*\d+(?:-\d+)?)\b",
        r"\b(a216\s*wcb)\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return re.sub(r"\s+", " ", match.group(1)).upper()
    return ""


def detect_material(name: str) -> str:
    text = normalize_text(name)
    if "паронит" in text or "пон-б" in text:
        return "Паронит"
    if "нерж" in text or "12х18" in text:
        return "Нержавеющая сталь"
    if "чугун" in text or re.search(r"\b30ч", text):
        return "Чугун"
    if "латун" in text:
        return "Латунь"
    if "бронз" in text:
        return "Бронза"
    if "болт" in text or "шпильк" in text or "гайк" in text:
        return "Сталь"
    if "стал" in text or "wcb" in text or re.search(r"\b30[сc]", text):
        return "Сталь"
    return "Не указан"


def detect_connection_type(name: str) -> str:
    text = normalize_text(name)
    if "болт" in text or "шпильк" in text or "гайк" in text:
        return "Резьбовое"
    if "межфланц" in text:
        return "Межфланцевое"
    if "под привар" in text or "привар" in text or "сварн" in text:
        return "Под приварку"
    if "lug" in text:
        return "Lug"
    if "фланц" in text:
        return "Фланцевое"
    if "муфт" in text:
        return "Муфтовое"
    return "Не указано"


def detect_control_type(name: str, category_id: str) -> str:
    text = normalize_text(name)
    if category_id == "elektroprivody":
        return "Электропривод"
    if "электроприв" in text:
        return "Электропривод"
    if "пневм" in text:
        return "Пневмопривод"
    if "редукт" in text:
        return "Редуктор"
    if category_id in {"flansy-i-otvody", "filtry-i-kompensatory", "krepezh-i-prokladki"}:
        return "Не требуется"
    if "обратн" in text and "клапан" in text:
        return "Автоматическое"
    return "Ручное"


def _classify_zadvizhki(text: str) -> tuple[str, str]:
    # Order matters:
    #   1. Shibernye / knife / ЗМС / Lug — dedicated subcategory regardless of material.
    #   2. Klinovye — dedicated subcategory regardless of material.
    #   3. Electroprivod — functional subcategory.
    #   4. High-pressure PN40/PN64.
    #   5. Fallback by material: chugun -> chugunnye, else stalnyye.
    if "шибер" in text or "ножев" in text or "змс" in text or "lug" in text:
        return "zadvizhki", "zadvizhki-shibernye"
    if "клинов" in text:
        return "zadvizhki", "zadvizhki-klinovye"
    if "электроприв" in text:
        return "zadvizhki", "zadvizhki-s-elektroprivodom"
    if "pn40" in text or "pn64" in text:
        return "zadvizhki", "zadvizhki-pn40-pn64"
    if "чугун" in text or "30ч" in text:
        return "zadvizhki", "zadvizhki-chugunnye"
    return "zadvizhki", "zadvizhki-stalnyye"


def _classify_krany(text: str) -> tuple[str, str]:
    # "цельносварн" wins even when "под приварку" also present in the name.
    if "цельносварн" in text:
        return "krany-sharovye", "krany-tselnosvarnye"
    if "привар" in text or "сварн" in text:
        return "krany-sharovye", "krany-pod-privarku"
    return "krany-sharovye", "krany-flantsevy"


def _classify_zatvory(text: str, sheet_text: str = "") -> tuple[str, str]:
    # Sheet name is an authoritative signal: the "Затвор уплотнение..." sheet
    # has product names that don't contain "уплотнен", so we rely on the sheet.
    if "уплотнен" in sheet_text or "уплотнен" in text:
        return "zatvory", "zatvory-s-uplotneniem"
    # "дисков" is a specific shape marker and wins over generic flange-type.
    if "дисков" in text:
        return "zatvory", "zatvory-diskovye"
    if "межфланц" in text:
        return "zatvory", "zatvory-mezhflantsevy"
    return "zatvory", "zatvory-flantsevy"


def _classify_klapany(text: str) -> tuple[str, str]:
    # Chugun + vodoprovod is a distinct product family, keep it first.
    # Then "обратн" must override "поворот" because many rows are named
    # "поворотный обратный клапан" and belong to "Обратные".
    if "чугун" in text and "водопровод" in text:
        return "klapany", "klapany-chugunnye"
    if "обратн" in text:
        return "klapany", "klapany-obratnye"
    if "поворот" in text:
        return "klapany", "klapany-povorotnye"
    return "klapany", "klapany-obratnye"


def _classify_elektroprivody(text: str) -> tuple[str, str]:
    if "задвиж" in text:
        return "elektroprivody", "elektroprivody-dlya-zadvizhek"
    if "затвор" in text:
        return "elektroprivody", "elektroprivody-dlya-zatvorov"
    if "клапан" in text:
        return "elektroprivody", "elektroprivody-dlya-klapanov"
    return "elektroprivody", "elektroprivody-dlya-zadvizhek"


def _classify_krepezh_i_prokladki(text: str) -> tuple[str, str]:
    if "болт" in text:
        return "krepezh-i-prokladki", "krepezh-bolty"
    if "шпильк" in text:
        return "krepezh-i-prokladki", "krepezh-shpilki"
    if "гайк" in text:
        return "krepezh-i-prokladki", "krepezh-gayki"
    if "паронит" in text or "пон-б" in text:
        return "krepezh-i-prokladki", "prokladki-paronit-ponb"
    return "krepezh-i-prokladki", "krepezh-bolty"


def _is_standalone_elektroprivod(text: str) -> bool:
    if "электропривод" not in text:
        return False
    return text.startswith("электропривод") or "для " in text


def category_from_sheet(sheet_name: str) -> tuple[str, str] | None:
    text = normalize_text(sheet_name)
    if "электропривод" in text:
        return _classify_elektroprivody(text)
    if any(marker in text for marker in ("болт", "шпильк", "гайк", "паронит", "пон-б")):
        return _classify_krepezh_i_prokladki(text)
    if "компенсатор" in text:
        return "filtry-i-kompensatory", "kompensatory"
    if "фильтр" in text or "корзинчат" in text:
        return "filtry-i-kompensatory", "filtry-korzinchatye"
    if "отвод" in text:
        return "flansy-i-otvody", "otvody"
    if "фланц" in text and "затвор" not in text and "задвиж" not in text and "кран" not in text:
        return "flansy-i-otvody", "flansy"
    if "кран" in text:
        return _classify_krany(text)
    if "затвор" in text:
        return _classify_zatvory(text, text)
    if "клапан" in text:
        return _classify_klapany(text)
    if "задвиж" in text or "шибер" in text:
        return _classify_zadvizhki(text)
    return None


def classify_category(product_name: str, sheet_name: str) -> tuple[str, str]:
    text = normalize_text(product_name)

    if _is_standalone_elektroprivod(text):
        return _classify_elektroprivody(text)
    if any(marker in text for marker in ("болт", "шпильк", "гайк", "паронит", "пон-б")):
        return _classify_krepezh_i_prokladki(text)

    if "компенсатор" in text:
        return "filtry-i-kompensatory", "kompensatory"
    if "фильтр" in text or "корзинчат" in text:
        return "filtry-i-kompensatory", "filtry-korzinchatye"
    if "отвод" in text:
        return "flansy-i-otvody", "otvody"
    if "фланец" in text and "затвор" not in text:
        return "flansy-i-otvody", "flansy"

    if "кран" in text:
        return _classify_krany(text)

    if "затвор" in text:
        return _classify_zatvory(text, normalize_text(sheet_name))

    if "клапан" in text:
        return _classify_klapany(text)

    if "задвиж" in text or "шибер" in text:
        return _classify_zadvizhki(text)

    fallback = category_from_sheet(sheet_name)
    if fallback:
        return fallback

    raise ValueError(f"Cannot map product to final structure: {product_name!r} (sheet={sheet_name!r})")


def build_seo_name(
    category_id: str,
    subcategory_id: str,
    model: str,
    material: str,
    dn: int | None,
    pn: int | None,
    thread: str | None,
) -> str:
    if category_id == "zadvizhki":
        base = "Задвижка"
    elif category_id == "krany-sharovye":
        base = "Кран шаровой"
    elif category_id == "zatvory":
        base = "Затвор"
    elif category_id == "klapany":
        base = "Клапан"
    elif category_id == "filtry-i-kompensatory":
        base = "Компенсатор" if subcategory_id == "kompensatory" else "Фильтр"
    elif category_id == "flansy-i-otvody":
        base = "Отвод" if subcategory_id == "otvody" else "Фланец"
    elif category_id == "krepezh-i-prokladki":
        if subcategory_id == "krepezh-bolty":
            base = "Болт фланцевый"
        elif subcategory_id == "krepezh-shpilki":
            base = "Шпилька фланцевая"
        elif subcategory_id == "krepezh-gayki":
            base = "Гайка фланцевая"
        else:
            base = "Прокладка паронитовая ПОН-Б"
    elif category_id == "elektroprivody":
        if subcategory_id == "elektroprivody-dlya-zatvorov":
            base = "Электропривод для затворов"
        elif subcategory_id == "elektroprivody-dlya-klapanov":
            base = "Электропривод для клапанов"
        else:
            base = "Электропривод для задвижек"
    else:
        base = "Арматура"

    parts = [base]
    if model:
        parts.append(model)
    elif material != "Не указан":
        parts.append(material.lower())

    if dn is not None:
        parts.append(f"DN{dn}")
    if pn is not None:
        parts.append(f"PN{pn}")
    if thread:
        parts.append(thread)

    return " ".join(parts)


def slugify(value: str) -> str:
    out: list[str] = []
    text = normalize_text(value)
    for ch in text:
        if ch in CYR_TO_LAT:
            out.append(CYR_TO_LAT[ch])
        elif "a" <= ch <= "z" or "0" <= ch <= "9":
            out.append(ch)
        else:
            out.append("-")
    slug = re.sub(r"-{2,}", "-", "".join(out)).strip("-")
    return slug or "product"


def format_number(value: float | None) -> float | None:
    if value is None:
        return None
    return round(value, 3)


@dataclass
class RawRow:
    source_file: str
    source_sheet: str
    source_row: int
    name_raw: str
    weight: float | None
    price: float | None


def header_index_map(header_row: list[Any]) -> dict[str, int]:
    result: dict[str, int] = {}
    for idx, value in enumerate(header_row):
        cell = normalize_text(str(value)) if value is not None else ""
        if not cell:
            continue
        if "товар" in cell:
            result["name"] = idx
        elif "вес" in cell:
            result["weight"] = idx
        elif "продажа" in cell:
            result["sale"] = idx
        elif "цена" in cell and "(₸" in str(value):
            result["price_kzt"] = idx
    return result


def parse_workbooks(excel_files: list[Path]) -> list[RawRow]:
    rows: list[RawRow] = []

    for file_path in excel_files:
        workbook = load_workbook(file_path, data_only=True)
        for sheet in workbook.worksheets:
            if normalize_text(sheet.title).startswith("лист"):
                continue

            values = list(sheet.iter_rows(values_only=True))
            if not values:
                continue

            headers = header_index_map(list(values[0]))
            name_idx = headers.get("name", 1)
            weight_idx = headers.get("weight", 5)
            sale_idx = headers.get("sale", 17)
            price_kzt_idx = headers.get("price_kzt", 4)

            for row_idx, row in enumerate(values[1:], start=2):
                name_raw = row[name_idx] if name_idx < len(row) else None
                if not isinstance(name_raw, str):
                    continue

                normalized_name = re.sub(r"\s+", " ", name_raw).strip()
                if len(normalized_name) < 5:
                    continue

                sale_price = to_number(row[sale_idx] if sale_idx < len(row) else None)
                fallback_price = to_number(row[price_kzt_idx] if price_kzt_idx < len(row) else None)
                price = sale_price if sale_price is not None else fallback_price

                if price is None:
                    # Section labels inside the sheet; skip.
                    continue

                weight = to_number(row[weight_idx] if weight_idx < len(row) else None)

                rows.append(
                    RawRow(
                        source_file=file_path.name,
                        source_sheet=sheet.title,
                        source_row=row_idx,
                        name_raw=normalized_name,
                        weight=weight,
                        price=price,
                    )
                )
    return rows


def load_override_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []

    payload = json.loads(path.read_text(encoding="utf-8"))
    raw_items = payload.get("products")
    if not isinstance(raw_items, list):
        raise ValueError("Invalid overrides format: expected top-level 'products' array")

    normalized: list[dict[str, Any]] = []
    for idx, raw_item in enumerate(raw_items, start=1):
        if not isinstance(raw_item, dict):
            raise ValueError(f"Invalid override product at index {idx}: expected object")

        category_id = str(raw_item.get("category", "")).strip()
        subcategory_id = str(raw_item.get("subcategory", "")).strip()
        if category_id not in CATEGORY_NAME_BY_ID:
            raise ValueError(f"Unknown override category: {category_id!r} at index {idx}")
        if subcategory_id not in SUBCATEGORY_BY_ID:
            raise ValueError(f"Unknown override subcategory: {subcategory_id!r} at index {idx}")

        name_raw = str(raw_item.get("nameRaw") or raw_item.get("name") or "").strip()
        if not name_raw:
            raise ValueError(f"Override product at index {idx} has no name/nameRaw")

        dn = (
            int(raw_item["dn"])
            if raw_item.get("dn") is not None and str(raw_item.get("dn")).strip() != ""
            else extract_dn(name_raw)
        )
        pn = (
            int(raw_item["pn"])
            if raw_item.get("pn") is not None and str(raw_item.get("pn")).strip() != ""
            else extract_pn(name_raw)
        )
        thread = str(raw_item["thread"]).strip().upper() if raw_item.get("thread") else extract_thread(name_raw)
        model = str(raw_item.get("model") or extract_model(name_raw)).strip()
        material = str(raw_item.get("material") or detect_material(name_raw)).strip()
        connection_type = str(raw_item.get("connectionType") or detect_connection_type(name_raw)).strip()
        control_type = str(raw_item.get("controlType") or detect_control_type(name_raw, category_id)).strip()

        name = str(raw_item.get("name") or "").strip()
        if not name:
            name = build_seo_name(category_id, subcategory_id, model, material, dn, pn, thread)

        price_raw = to_number(raw_item.get("price"))
        price = format_number(price_raw)
        price_by_request = bool(raw_item.get("priceByRequest", price is None or (price is not None and price <= 0)))
        if price_by_request:
            price = None

        weight = format_number(to_number(raw_item.get("weight")))
        specs = raw_item.get("specs") if isinstance(raw_item.get("specs"), dict) else {}
        specs = {str(k): str(v) for k, v in specs.items()}
        if dn is not None and "DN" not in specs:
            specs["DN"] = str(dn)
        if pn is not None and "PN" not in specs:
            specs["PN"] = str(pn)
        if thread and "Резьба" not in specs:
            specs["Резьба"] = thread
        if material != "Не указан" and "Материал" not in specs:
            specs["Материал"] = material
        if connection_type != "Не указано" and "Присоединение" not in specs:
            specs["Присоединение"] = connection_type
        if control_type not in {"", "Не требуется"} and "Тип управления" not in specs:
            specs["Тип управления"] = control_type
        if model and "Модель" not in specs:
            specs["Модель"] = model
        if weight is not None and "Вес, кг" not in specs:
            specs["Вес, кг"] = str(weight)

        short_description = str(raw_item.get("shortDescription") or "").strip()
        if not short_description:
            short_desc_parts = [name, f"Категория: {CATEGORY_NAME_BY_ID[category_id]}"]
            if material != "Не указан":
                short_desc_parts.append(f"материал {material.lower()}")
            if thread:
                short_desc_parts.append(f"резьба {thread}")
            short_description = ". ".join(short_desc_parts) + "."

        normalized.append(
            {
                "name": name,
                "nameRaw": name_raw,
                "category": category_id,
                "subcategory": subcategory_id,
                "subcategoryName": SUBCATEGORY_BY_ID[subcategory_id]["name"],
                "categoryName": CATEGORY_NAME_BY_ID[category_id],
                "dn": dn,
                "pn": pn,
                "thread": thread,
                "material": material,
                "connectionType": connection_type,
                "controlType": control_type,
                "model": model,
                "price": price if not price_by_request else None,
                "priceByRequest": price_by_request,
                "weight": weight,
                "specs": specs,
                "shortDescription": short_description,
                "sourceFile": path.name,
                "sourceSheet": "__overrides__",
                "sourceRow": idx,
            }
        )

    return normalized


def dedupe_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    def source_priority(item: dict[str, Any]) -> int:
        file_name = normalize_text(item["sourceFile"])
        if "override" in file_name:
            return 3
        return 2 if "втор" in file_name else 1

    index: dict[str, dict[str, Any]] = {}
    for item in rows:
        dedupe_key = "|".join(
            [
                normalize_text(item["nameRaw"]),
                item["category"],
                item["subcategory"],
                str(item["dn"] or ""),
                str(item["pn"] or ""),
                str(item.get("thread") or ""),
            ]
        )
        existing = index.get(dedupe_key)
        if not existing:
            index[dedupe_key] = item
            continue

        keep_new = False
        if source_priority(item) > source_priority(existing):
            keep_new = True
        elif (item["price"] is not None) and (existing["price"] is None):
            keep_new = True
        elif (item["price"] or 0) > (existing["price"] or 0):
            keep_new = True
        elif len(item["nameRaw"]) > len(existing["nameRaw"]):
            keep_new = True

        if keep_new:
            index[dedupe_key] = item

    return list(index.values())


def build_products(
    rows: list[RawRow], override_rows: list[dict[str, Any]] | None = None
) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []

    for row in rows:
        dn = extract_dn(row.name_raw)
        pn = extract_pn(row.name_raw)
        thread = extract_thread(row.name_raw)
        category_id, subcategory_id = classify_category(row.name_raw, row.source_sheet)
        model = extract_model(row.name_raw)
        material = detect_material(row.name_raw)
        connection_type = detect_connection_type(row.name_raw)
        control_type = detect_control_type(row.name_raw, category_id)
        name = build_seo_name(category_id, subcategory_id, model, material, dn, pn, thread)
        price = format_number(row.price)
        weight = format_number(row.weight)
        price_by_request = price is None or price <= 0

        specs: dict[str, str] = {}
        if dn is not None:
            specs["DN"] = str(dn)
        if pn is not None:
            specs["PN"] = str(pn)
        if thread:
            specs["Резьба"] = thread
        if material != "Не указан":
            specs["Материал"] = material
        if connection_type != "Не указано":
            specs["Присоединение"] = connection_type
        if control_type not in {"", "Не требуется"}:
            specs["Тип управления"] = control_type
        if model:
            specs["Модель"] = model
        if weight is not None:
            specs["Вес, кг"] = str(weight)

        short_desc_parts = [name, f"Категория: {CATEGORY_NAME_BY_ID[category_id]}"]
        if material != "Не указан":
            short_desc_parts.append(f"материал {material.lower()}")
        if thread:
            short_desc_parts.append(f"резьба {thread}")
        if connection_type != "Не указано":
            short_desc_parts.append(f"присоединение {connection_type.lower()}")
        short_description = ". ".join(short_desc_parts) + "."

        normalized.append(
            {
                "name": name,
                "nameRaw": row.name_raw,
                "category": category_id,
                "subcategory": subcategory_id,
                "subcategoryName": SUBCATEGORY_BY_ID[subcategory_id]["name"],
                "categoryName": CATEGORY_NAME_BY_ID[category_id],
                "dn": dn,
                "pn": pn,
                "thread": thread,
                "material": material,
                "connectionType": connection_type,
                "controlType": control_type,
                "model": model,
                "price": price if not price_by_request else None,
                "priceByRequest": price_by_request,
                "weight": weight,
                "specs": specs,
                "shortDescription": short_description,
                "sourceFile": row.source_file,
                "sourceSheet": row.source_sheet,
                "sourceRow": row.source_row,
            }
        )

    if override_rows:
        normalized.extend(override_rows)

    deduped = dedupe_rows(normalized)
    deduped.sort(
        key=lambda item: (
            item["category"],
            item["subcategory"],
            item["dn"] if item["dn"] is not None else 99999,
            item["pn"] if item["pn"] is not None else 999,
            item["name"],
        )
    )

    slug_counts: dict[str, int] = {}
    products: list[dict[str, Any]] = []
    for idx, item in enumerate(deduped, start=1):
        base_slug = slugify(item["name"])
        occurrence = slug_counts.get(base_slug, 0) + 1
        slug_counts[base_slug] = occurrence
        slug = base_slug if occurrence == 1 else f"{base_slug}-{occurrence}"

        products.append(
            {
                "id": str(idx),
                "name": item["name"],
                "slug": slug,
                "category": item["category"],
                "subcategory": item["subcategory"],
                "subcategoryName": item["subcategoryName"],
                "categoryName": item["categoryName"],
                "dn": item["dn"],
                "pn": item["pn"],
                "thread": item.get("thread"),
                "material": item["material"],
                "connectionType": item["connectionType"],
                "controlType": item["controlType"],
                "model": item["model"],
                "price": item["price"],
                "priceByRequest": item["priceByRequest"],
                "weight": item["weight"],
                "specs": item["specs"],
                "shortDescription": item["shortDescription"],
            }
        )

    return products


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_ts(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    content = """// Auto-generated from Excel files.
// DO NOT EDIT MANUALLY — run `python scripts/rebuild_catalog.py`

import catalogData from "@/data/catalog-products.json";

export interface Product {
  id: string;
  name: string;
  slug: string;
  category: string;
  subcategory: string;
  subcategoryName: string;
  categoryName: string;
  dn: number | undefined;
  pn: number | undefined;
  thread: string | undefined;
  material: string;
  connectionType: string;
  controlType: string;
  model: string;
  price: number | undefined;
  priceByRequest: boolean;
  weight: number | undefined;
  specs: Record<string, string>;
  shortDescription: string;
}

export interface Category {
  id: string;
  name: string;
  slug: string;
  subcategories: Subcategory[];
}

export interface Subcategory {
  id: string;
  name: string;
  slug: string;
  parentCategory: string;
}

type RawProduct = Omit<Product, "dn" | "pn" | "thread" | "price" | "weight"> & {
  dn: number | null;
  pn: number | null;
  thread: string | null;
  price: number | null;
  weight: number | null;
};

const rawCategories = (catalogData.categories ?? []) as Category[];
const rawProducts = (catalogData.products ?? []) as unknown as RawProduct[];

export const categories: Category[] = rawCategories;

export const products: Product[] = rawProducts.map((product) => ({
  ...product,
  dn: product.dn ?? undefined,
  pn: product.pn ?? undefined,
  thread: product.thread ?? undefined,
  price: product.price ?? undefined,
  weight: product.weight ?? undefined,
}));

export function getProductsByCategory(categoryId: string): Product[] {
  return products.filter((p) => p.category === categoryId);
}

export function getProductsBySubcategory(subcategoryId: string): Product[] {
  return products.filter((p) => p.subcategory === subcategoryId);
}

export function getProductBySlug(slug: string): Product | undefined {
  return products.find((p) => p.slug === slug);
}

export function getCategoryById(id: string): Category | undefined {
  return categories.find((c) => c.id === id);
}

export function getCategoryBySlug(slug: string): Category | undefined {
  return categories.find((c) => c.slug === slug);
}

export function searchProducts(query: string): Product[] {
  const q = query.toLowerCase().trim();
  if (!q) return products;

  return products.filter((p) => {
    const haystack = [
      p.name,
      p.material,
      p.connectionType,
      p.controlType,
      p.model,
      p.categoryName,
      p.subcategoryName,
      p.shortDescription,
      p.dn != null ? String(p.dn) : "",
      p.pn != null ? String(p.pn) : "",
      p.thread ?? "",
    ]
      .join(" ")
      .toLowerCase();

    return haystack.includes(q);
  });
}
"""
    path.write_text(content, encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Rebuild catalog from Excel files")
    parser.add_argument(
        "--excel-file",
        type=Path,
        action="append",
        default=[],
        help="Explicit Excel file path (can be repeated)",
    )
    parser.add_argument(
        "--excel-dir",
        type=Path,
        default=Path(r"C:\Users\user\Desktop\Work\Черновик\Mansvalve"),
        help="Directory containing source .xlsx files (used when --excel-file is not set)",
    )
    parser.add_argument(
        "--output-json",
        type=Path,
        default=Path("data/catalog-products.json"),
        help="Path to normalized JSON output",
    )
    parser.add_argument(
        "--output-ts",
        type=Path,
        default=Path("lib/catalog-data.ts"),
        help="Path to generated TypeScript module",
    )
    parser.add_argument(
        "--overrides-json",
        type=Path,
        default=Path("data/catalog-overrides.json"),
        help="Optional JSON with manual product overrides",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.excel_file:
        excel_files = [path for path in args.excel_file if path.suffix.lower() == ".xlsx"]
    else:
        excel_files = sorted(args.excel_dir.glob("MANSVALVE*.xlsx"))

    if not excel_files:
        raise SystemExit(
            "No input .xlsx files found. Use --excel-file or place MANSVALVE*.xlsx in --excel-dir."
        )

    raw_rows = parse_workbooks(excel_files)
    if not raw_rows:
        raise SystemExit("No product rows found in Excel files")

    override_rows = load_override_rows(args.overrides_json)
    products = build_products(raw_rows, override_rows=override_rows)
    payload = {
        "meta": {
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "sourceFiles": [p.name for p in excel_files],
            "rawRows": len(raw_rows),
            "overrideProducts": len(override_rows),
            "products": len(products),
        },
        "categories": CATEGORIES,
        "products": products,
    }

    write_json(args.output_json, payload)
    write_ts(args.output_ts)

    print(f"Generated {args.output_json} and {args.output_ts}")
    print(
        f"Source files: {len(excel_files)} | raw rows: {len(raw_rows)} | "
        f"overrides: {len(override_rows)} | products: {len(products)}"
    )


if __name__ == "__main__":
    main()
